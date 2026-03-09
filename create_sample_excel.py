#!/usr/bin/env python3
"""Creates a sample scripts.xlsx for Script Runner."""
import subprocess, sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Scripts"

# Header
headers = ["Category", "Action", "Script Path"]
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2d2d44", end_color="2d2d44", fill_type="solid")
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Sample data
data = [
    ("System", "Check Disk Usage", "/usr/local/scripts/disk_usage.sh"),
    ("System", "Check Memory", "/usr/local/scripts/memory_check.sh"),
    ("System", "List Services", "/usr/local/scripts/list_services.sh"),
    ("System", "System Info", "/usr/local/scripts/sysinfo.sh"),
    ("Network", "Ping Test", "/usr/local/scripts/ping_test.sh"),
    ("Network", "Port Scan", "/usr/local/scripts/port_scan.sh"),
    ("Network", "Check DNS", "/usr/local/scripts/dns_check.sh"),
    ("Backup", "Backup Database", "/usr/local/scripts/backup_db.sh"),
    ("Backup", "Backup Files", "/usr/local/scripts/backup_files.sh"),
    ("Backup", "Restore Latest", "/usr/local/scripts/restore.sh"),
    ("Deploy", "Deploy Staging", "/usr/local/scripts/deploy_staging.sh"),
    ("Deploy", "Deploy Production", "/usr/local/scripts/deploy_prod.sh"),
    ("Deploy", "Rollback", "/usr/local/scripts/rollback.sh"),
    ("Maintenance", "Clear Logs", "/usr/local/scripts/clear_logs.sh"),
    ("Maintenance", "Update Packages", "/usr/local/scripts/update_packages.sh"),
    ("Maintenance", "Restart Services", "/usr/local/scripts/restart_all.sh"),
]

for row, (cat, action, path) in enumerate(data, 2):
    ws.cell(row=row, column=1, value=cat)
    ws.cell(row=row, column=2, value=action)
    ws.cell(row=row, column=3, value=path)

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 45

wb.save("scripts.xlsx")
print("✓ Created scripts.xlsx with sample data")
