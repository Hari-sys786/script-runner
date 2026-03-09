#!/usr/bin/env python3
"""
Server Manager v2.0 — Web Edition
Bank → Server → Application → Refresh/Restart
"""

import os
import sys
import json
import subprocess
import threading
import time

try:
    from flask import Flask, render_template_string, request, jsonify, redirect, url_for, session as flask_session
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "flask", "--break-system-packages"])
    from flask import Flask, render_template_string, request, jsonify, redirect, url_for, session as flask_session

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
    from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'servermanager_v2_secret_key_2026'

ADMIN_USER = 'admin'
ADMIN_PASS = 'admin123'

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "servers.xlsx")

running_processes = {}
process_lock = threading.Lock()


def load_excel():
    """Returns: {bank: {server: {app: script_path}}}"""
    data = {}
    if not os.path.exists(EXCEL_FILE):
        return data
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4 or not row[0]:
            continue
        bank = str(row[0]).strip()
        server = str(row[1]).strip()
        application = str(row[2]).strip()
        script = str(row[3]).strip()
        if bank not in data:
            data[bank] = {}
        if server not in data[bank]:
            data[bank][server] = {}
        data[bank][server][application] = script
    wb.close()
    return data


LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>⚡ Server Manager — Login</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            background: #1e1e2e;
            color: #cdd6f4;
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .login-card {
            background: #2a2a3c;
            border-radius: 16px;
            padding: 40px;
            width: 380px;
            box-shadow: 0 8px 32px rgba(0,0,0,0.3);
        }
        .login-card .icon {
            text-align: center;
            font-size: 48px;
            margin-bottom: 8px;
        }
        .login-card h1 {
            text-align: center;
            font-size: 22px;
            color: #89b4fa;
            margin-bottom: 4px;
        }
        .login-card .subtitle {
            text-align: center;
            color: #6c7086;
            font-size: 13px;
            margin-bottom: 28px;
        }
        .form-group {
            margin-bottom: 16px;
        }
        .form-group label {
            display: block;
            font-size: 12px;
            color: #6c7086;
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            font-weight: 600;
        }
        .form-group input {
            width: 100%;
            padding: 11px 14px;
            background: #313244;
            color: #cdd6f4;
            border: 1px solid #45475a;
            border-radius: 8px;
            font-size: 14px;
            font-family: inherit;
            transition: border-color 0.2s;
        }
        .form-group input:focus {
            outline: none;
            border-color: #89b4fa;
            box-shadow: 0 0 0 2px rgba(137,180,250,0.2);
        }
        .form-group input::placeholder {
            color: #585b70;
        }
        .btn-login {
            width: 100%;
            padding: 12px;
            background: #89b4fa;
            color: #1e1e2e;
            border: none;
            border-radius: 8px;
            font-size: 15px;
            font-weight: 700;
            cursor: pointer;
            transition: all 0.2s;
            font-family: inherit;
            margin-top: 8px;
        }
        .btn-login:hover { background: #74c7ec; }
        .btn-login:active { transform: scale(0.98); }
        .error-msg {
            background: rgba(243,139,168,0.15);
            border: 1px solid #f38ba8;
            color: #f38ba8;
            padding: 10px 14px;
            border-radius: 8px;
            font-size: 13px;
            margin-bottom: 16px;
            text-align: center;
        }
        .footer {
            text-align: center;
            margin-top: 20px;
            color: #45475a;
            font-size: 12px;
        }
    </style>
</head>
<body>
    <div class="login-card">
        <div class="icon">⚡</div>
        <h1>Server Manager</h1>
        <p class="subtitle">Sign in to continue</p>
        {% if error %}
        <div class="error-msg">{{ error }}</div>
        {% endif %}
        <form method="POST" action="/login">
            <div class="form-group">
                <label>Username</label>
                <input type="text" name="username" placeholder="Enter username" autofocus required>
            </div>
            <div class="form-group">
                <label>Password</label>
                <input type="password" name="password" placeholder="Enter password" required>
            </div>
            <button type="submit" class="btn-login">Sign In</button>
        </form>
        <div class="footer">Server Manager v2.0</div>
    </div>
</body>
</html>
"""

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>⚡ Server Manager v2.0</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }

        :root {
            --bg: #1e1e2e;
            --bg-secondary: #2a2a3c;
            --fg: #cdd6f4;
            --fg-dim: #6c7086;
            --accent: #89b4fa;
            --accent-hover: #74c7ec;
            --success: #a6e3a1;
            --error: #f38ba8;
            --warning: #fab387;
            --border: #45475a;
            --input-bg: #313244;
            --button-bg: #585b70;
            --refresh: #a6e3a1;
            --refresh-hover: #94e2d5;
            --restart: #fab387;
            --restart-hover: #f9e2af;
        }

        body {
            background: var(--bg);
            color: var(--fg);
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
            min-height: 100vh;
            padding: 30px;
        }

        .container { max-width: 850px; margin: 0 auto; }

        .header h1 {
            font-size: 28px;
            color: var(--accent);
            margin-bottom: 4px;
        }
        .header p {
            color: var(--fg-dim);
            font-size: 14px;
            margin-bottom: 20px;
        }

        .card {
            background: var(--bg-secondary);
            border-radius: 12px;
            padding: 24px;
            margin-bottom: 16px;
        }

        .form-row {
            display: flex;
            gap: 16px;
            margin-bottom: 12px;
        }

        .form-group { flex: 1; }

        .form-group label {
            display: block;
            font-size: 12px;
            color: var(--fg-dim);
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.8px;
            font-weight: 600;
        }

        select {
            width: 100%;
            padding: 10px 14px;
            background: var(--input-bg);
            color: var(--fg);
            border: 1px solid var(--border);
            border-radius: 8px;
            font-size: 14px;
            font-family: inherit;
            cursor: pointer;
            appearance: none;
            -webkit-appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%2389b4fa' d='M6 8L1 3h10z'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 12px center;
            transition: border-color 0.2s;
        }

        select:focus {
            outline: none;
            border-color: var(--accent);
            box-shadow: 0 0 0 2px rgba(137, 180, 250, 0.2);
        }

        select option {
            background: var(--input-bg);
            color: var(--fg);
        }

        .script-path {
            font-family: 'Consolas', 'Fira Code', monospace;
            font-size: 12px;
            color: var(--fg-dim);
            padding: 8px 0 0;
            min-height: 20px;
        }

        .btn-row {
            display: flex;
            gap: 10px;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }

        .btn {
            padding: 10px 24px;
            border: none;
            border-radius: 8px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
            font-family: inherit;
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

        .btn:active { transform: scale(0.97); }

        .btn-refresh {
            background: var(--refresh);
            color: var(--bg);
        }
        .btn-refresh:hover { background: var(--refresh-hover); }
        .btn-refresh:disabled {
            background: var(--border);
            color: var(--fg-dim);
            cursor: not-allowed;
        }

        .btn-restart {
            background: var(--restart);
            color: var(--bg);
        }
        .btn-restart:hover { background: var(--restart-hover); }
        .btn-restart:disabled {
            background: var(--border);
            color: var(--fg-dim);
            cursor: not-allowed;
        }

        .btn-kill {
            background: var(--error);
            color: var(--bg);
        }
        .btn-kill:hover { background: #eba0ac; }
        .btn-kill:disabled {
            background: var(--border);
            color: var(--fg-dim);
            cursor: not-allowed;
        }

        .btn-secondary {
            background: var(--button-bg);
            color: var(--fg);
        }
        .btn-secondary:hover { background: var(--border); }

        .btn-row .spacer { flex: 1; }

        .output-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 8px;
        }
        .output-label {
            font-size: 12px;
            color: var(--fg-dim);
            text-transform: uppercase;
            letter-spacing: 0.8px;
            font-weight: 600;
        }

        #output {
            background: var(--input-bg);
            border: 1px solid var(--border);
            border-radius: 8px;
            padding: 16px;
            font-family: 'Consolas', 'Fira Code', monospace;
            font-size: 13px;
            line-height: 1.6;
            min-height: 280px;
            max-height: 450px;
            overflow-y: auto;
            white-space: pre-wrap;
            word-wrap: break-word;
        }

        #output::-webkit-scrollbar { width: 6px; }
        #output::-webkit-scrollbar-track { background: var(--input-bg); }
        #output::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }

        .status-bar {
            margin-top: 12px;
            font-size: 13px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .status-dot {
            width: 8px;
            height: 8px;
            border-radius: 50%;
            display: inline-block;
        }
        .status-dot.ready { background: var(--success); }
        .status-dot.running { background: var(--warning); animation: pulse 1s infinite; }
        .status-dot.error { background: var(--error); }

        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.4; }
        }

        .text-success { color: var(--success); }
        .text-error { color: var(--error); }
        .text-warning { color: var(--warning); }
        .text-accent { color: var(--accent); }

        .info-badge {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            background: var(--input-bg);
            border: 1px solid var(--border);
            border-radius: 6px;
            padding: 4px 10px;
            font-size: 12px;
            color: var(--fg-dim);
        }
        .info-badge .dot {
            width: 6px; height: 6px;
            border-radius: 50%;
            background: var(--accent);
        }

        @media (max-width: 700px) {
            body { padding: 16px; }
            .form-row { flex-direction: column; }
            .btn-row { flex-direction: column; }
            .btn { width: 100%; justify-content: center; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header" style="display:flex; justify-content:space-between; align-items:flex-start;">
            <div>
                <h1>⚡ Server Manager</h1>
                <p>Bank → Server → Application → Refresh / Restart</p>
            </div>
            <a href="/logout" class="btn btn-secondary" style="text-decoration:none; margin-top:4px; font-size:13px;">🚪 Logout</a>
        </div>

        <div class="card">
            <div class="form-row">
                <div class="form-group">
                    <label>🏦 Bank Name</label>
                    <select id="bank" onchange="onBankChange()">
                        <option value="">-- Select Bank --</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>🖥️ Server Name</label>
                    <select id="server" onchange="onServerChange()">
                        <option value="">-- Select Server --</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>📦 Application</label>
                    <select id="application" onchange="onAppChange()">
                        <option value="">-- Select Application --</option>
                    </select>
                </div>
            </div>
            <div class="script-path" id="scriptPath"></div>
        </div>

        <div class="btn-row">
            <button class="btn btn-refresh" id="refreshBtn" onclick="runAction('refresh')">🔄  Refresh</button>
            <button class="btn btn-restart" id="restartBtn" onclick="runAction('restart')">⚡  Restart</button>
            <button class="btn btn-kill" id="killBtn" onclick="killScript()" disabled>■  Stop</button>
            <div class="spacer"></div>
            <button class="btn btn-secondary" onclick="clearOutput()">Clear</button>
            <button class="btn btn-secondary" onclick="reloadExcel()">↻  Reload Excel</button>
        </div>

        <div class="output-header">
            <span class="output-label">Output</span>
            <span class="info-badge" id="infoBadge" style="display:none">
                <span class="dot"></span>
                <span id="infoText"></span>
            </span>
        </div>
        <div id="output"></div>

        <div class="status-bar">
            <span class="status-dot ready" id="statusDot"></span>
            <span id="statusText">Ready</span>
        </div>
    </div>

    <script>
        let scriptData = {};
        let currentProcessId = null;

        fetch('/api/data').then(r => r.json()).then(data => {
            scriptData = data;
            const bankSelect = document.getElementById('bank');
            Object.keys(data).sort().forEach(bank => {
                const opt = document.createElement('option');
                opt.value = bank;
                opt.textContent = bank;
                bankSelect.appendChild(opt);
            });
            let total = 0;
            Object.values(data).forEach(servers => {
                Object.values(servers).forEach(apps => {
                    total += Object.keys(apps).length;
                });
            });
            setStatus('ready', 'Loaded ' + Object.keys(data).length + ' banks, ' + total + ' entries');
        });

        function onBankChange() {
            const bank = document.getElementById('bank').value;
            const serverSelect = document.getElementById('server');
            const appSelect = document.getElementById('application');
            serverSelect.innerHTML = '<option value="">-- Select Server --</option>';
            appSelect.innerHTML = '<option value="">-- Select Application --</option>';
            document.getElementById('scriptPath').textContent = '';
            hideInfo();

            if (bank && scriptData[bank]) {
                Object.keys(scriptData[bank]).sort().forEach(srv => {
                    const opt = document.createElement('option');
                    opt.value = srv;
                    opt.textContent = srv;
                    serverSelect.appendChild(opt);
                });
            }
        }

        function onServerChange() {
            const bank = document.getElementById('bank').value;
            const server = document.getElementById('server').value;
            const appSelect = document.getElementById('application');
            appSelect.innerHTML = '<option value="">-- Select Application --</option>';
            document.getElementById('scriptPath').textContent = '';
            hideInfo();

            if (bank && server && scriptData[bank] && scriptData[bank][server]) {
                Object.keys(scriptData[bank][server]).sort().forEach(appName => {
                    const opt = document.createElement('option');
                    opt.value = appName;
                    opt.textContent = appName;
                    appSelect.appendChild(opt);
                });
            }
        }

        function onAppChange() {
            const bank = document.getElementById('bank').value;
            const server = document.getElementById('server').value;
            const application = document.getElementById('application').value;
            if (bank && server && application && scriptData[bank]?.[server]?.[application]) {
                document.getElementById('scriptPath').textContent = '📄 ' + scriptData[bank][server][application];
                showInfo(bank + ' / ' + server + ' / ' + application);
            } else {
                document.getElementById('scriptPath').textContent = '';
                hideInfo();
            }
        }

        function showInfo(text) {
            document.getElementById('infoBadge').style.display = 'inline-flex';
            document.getElementById('infoText').textContent = text;
        }
        function hideInfo() {
            document.getElementById('infoBadge').style.display = 'none';
        }

        function runAction(actionType) {
            const bank = document.getElementById('bank').value;
            const server = document.getElementById('server').value;
            const application = document.getElementById('application').value;

            if (!bank || !server || !application) {
                alert('Select Bank, Server, and Application first.');
                return;
            }

            const scriptPath = scriptData[bank]?.[server]?.[application];
            const label = actionType === 'refresh' ? '🔄 REFRESH' : '⚡ RESTART';

            appendOutput('\\n' + '─'.repeat(55) + '\\n', 'accent');
            appendOutput(label + ': ' + application + ' on ' + server + ' (' + bank + ')\\n', 'accent');
            appendOutput('Script: ' + scriptPath + '\\n', 'accent');
            appendOutput('─'.repeat(55) + '\\n', 'accent');

            setStatus('running', actionType.charAt(0).toUpperCase() + actionType.slice(1) + 'ing ' + application + '...');
            document.getElementById('refreshBtn').disabled = true;
            document.getElementById('restartBtn').disabled = true;
            document.getElementById('killBtn').disabled = false;

            fetch('/api/execute', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({bank, server, application, action: actionType})
            }).then(r => r.json()).then(data => {
                if (data.error) {
                    appendOutput('✗ ' + data.error + '\\n', 'error');
                    setStatus('error', data.error);
                    resetButtons();
                    return;
                }
                currentProcessId = data.process_id;
                pollOutput(actionType);
            });
        }

        function pollOutput(actionType) {
            if (!currentProcessId) return;

            fetch('/api/output/' + currentProcessId).then(r => r.json()).then(data => {
                if (data.lines && data.lines.length > 0) {
                    data.lines.forEach(line => appendOutput(line));
                }

                if (data.running) {
                    setTimeout(() => pollOutput(actionType), 300);
                } else {
                    const label = actionType === 'refresh' ? 'Refresh' : 'Restart';
                    if (data.exit_code === 0) {
                        appendOutput('\\n✓ ' + label + ' completed (exit 0)\\n', 'success');
                        setStatus('ready', label + ' of ' + data.application + ' completed');
                    } else if (data.exit_code === -9 || data.exit_code === -15) {
                        appendOutput('\\n■ Stopped by user\\n', 'warning');
                        setStatus('ready', 'Stopped');
                    } else {
                        appendOutput('\\n✗ ' + label + ' failed (exit ' + data.exit_code + ')\\n', 'error');
                        setStatus('error', label + ' failed (exit ' + data.exit_code + ')');
                    }
                    currentProcessId = null;
                    resetButtons();
                }
            });
        }

        function killScript() {
            if (currentProcessId) {
                fetch('/api/kill/' + currentProcessId, {method: 'POST'});
            }
        }

        function resetButtons() {
            document.getElementById('refreshBtn').disabled = false;
            document.getElementById('restartBtn').disabled = false;
            document.getElementById('killBtn').disabled = true;
        }

        function clearOutput() {
            document.getElementById('output').innerHTML = '';
        }

        function reloadExcel() {
            fetch('/api/reload').then(r => r.json()).then(data => {
                scriptData = data.data;
                document.getElementById('bank').innerHTML = '<option value="">-- Select Bank --</option>';
                document.getElementById('server').innerHTML = '<option value="">-- Select Server --</option>';
                document.getElementById('application').innerHTML = '<option value="">-- Select Application --</option>';
                document.getElementById('scriptPath').textContent = '';
                hideInfo();
                Object.keys(data.data).sort().forEach(bank => {
                    const opt = document.createElement('option');
                    opt.value = bank;
                    opt.textContent = bank;
                    document.getElementById('bank').appendChild(opt);
                });
                setStatus('ready', 'Reloaded: ' + data.banks + ' banks, ' + data.total + ' entries');
            });
        }

        function appendOutput(text, color) {
            const output = document.getElementById('output');
            const span = document.createElement('span');
            span.textContent = text;
            if (color) span.className = 'text-' + color;
            output.appendChild(span);
            output.scrollTop = output.scrollHeight;
        }

        function setStatus(type, text) {
            document.getElementById('statusDot').className = 'status-dot ' + type;
            const st = document.getElementById('statusText');
            st.textContent = text;
            st.className = type === 'error' ? 'text-error' : type === 'running' ? 'text-warning' : 'text-success';
        }
    </script>
</body>
</html>
"""


@app.route('/')
def index():
    if not flask_session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template_string(HTML_TEMPLATE)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if flask_session.get('logged_in'):
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        if username == ADMIN_USER and password == ADMIN_PASS:
            flask_session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            error = 'Invalid username or password'
    return render_template_string(LOGIN_TEMPLATE, error=error)


@app.route('/logout')
def logout():
    flask_session.clear()
    return redirect(url_for('login'))


@app.route('/api/data')
def get_data():
    if not flask_session.get('logged_in'):
        return jsonify({"error": "unauthorized"}), 401
    return jsonify(load_excel())


@app.route('/api/reload')
def reload_data():
    data = load_excel()
    total = sum(len(apps) for servers in data.values() for apps in servers.values())
    return jsonify({"data": data, "banks": len(data), "total": total})


@app.route('/api/execute', methods=['POST'])
def execute():
    req = request.json
    bank = req.get('bank', '')
    server = req.get('server', '')
    application = req.get('application', '')
    action = req.get('action', 'refresh')
    data = load_excel()

    if bank not in data or server not in data.get(bank, {}) or application not in data.get(bank, {}).get(server, {}):
        return jsonify({"error": "Invalid selection"})

    script_path = data[bank][server][application]
    if not os.path.exists(script_path):
        return jsonify({"error": f"Script not found: {script_path}"})

    pid = str(int(time.time() * 1000))

    def run_process():
        try:
            env = os.environ.copy()
            env['ACTION'] = action
            env['BANK'] = bank
            env['SERVER'] = server
            env['APPLICATION'] = application

            proc = subprocess.Popen(
                ["bash", script_path, action],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, env=env
            )
            with process_lock:
                running_processes[pid] = {
                    "process": proc, "output": [], "read_index": 0,
                    "running": True, "exit_code": None,
                    "application": application, "action": action
                }

            for line in proc.stdout:
                with process_lock:
                    if pid in running_processes:
                        running_processes[pid]["output"].append(line)

            proc.wait()
            with process_lock:
                if pid in running_processes:
                    running_processes[pid]["running"] = False
                    running_processes[pid]["exit_code"] = proc.returncode
        except Exception as e:
            with process_lock:
                if pid in running_processes:
                    running_processes[pid]["output"].append(f"Error: {e}\n")
                    running_processes[pid]["running"] = False
                    running_processes[pid]["exit_code"] = 1

    threading.Thread(target=run_process, daemon=True).start()
    time.sleep(0.1)
    return jsonify({"process_id": pid})


@app.route('/api/output/<pid>')
def get_output(pid):
    with process_lock:
        proc_info = running_processes.get(pid)
        if not proc_info:
            return jsonify({"error": "Process not found", "running": False})

        idx = proc_info["read_index"]
        new_lines = proc_info["output"][idx:]
        proc_info["read_index"] = len(proc_info["output"])

        return jsonify({
            "lines": new_lines, "running": proc_info["running"],
            "exit_code": proc_info["exit_code"],
            "application": proc_info["application"],
            "action": proc_info["action"]
        })


@app.route('/api/kill/<pid>', methods=['POST'])
def kill_process(pid):
    with process_lock:
        proc_info = running_processes.get(pid)
        if proc_info and proc_info["running"]:
            proc_info["process"].kill()
            return jsonify({"status": "killed"})
    return jsonify({"status": "not found"})


if __name__ == '__main__':
    print("⚡ Server Manager v2.0 — http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=False)
