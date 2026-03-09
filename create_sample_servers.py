#!/usr/bin/env python3
"""Creates sample servers.xlsx for Server Manager v2."""
import subprocess, sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Servers"

headers = ["Bank Name", "Server Name", "Application", "Script Path"]
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2d2d44", end_color="2d2d44", fill_type="solid")
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

data = [
    ("HDFC Bank", "HDFC-PROD-01", "Core Banking", "/usr/local/scripts/hdfc/core_banking.sh"),
    ("HDFC Bank", "HDFC-PROD-01", "Payment Gateway", "/usr/local/scripts/hdfc/payment_gw.sh"),
    ("HDFC Bank", "HDFC-PROD-01", "API Server", "/usr/local/scripts/hdfc/api_server.sh"),
    ("HDFC Bank", "HDFC-PROD-02", "Mobile Banking", "/usr/local/scripts/hdfc/mobile_banking.sh"),
    ("HDFC Bank", "HDFC-PROD-02", "Notification Service", "/usr/local/scripts/hdfc/notifications.sh"),
    ("HDFC Bank", "HDFC-UAT-01", "Core Banking UAT", "/usr/local/scripts/hdfc/core_banking_uat.sh"),
    ("SBI Bank", "SBI-PROD-01", "YONO App", "/usr/local/scripts/sbi/yono.sh"),
    ("SBI Bank", "SBI-PROD-01", "NEFT Service", "/usr/local/scripts/sbi/neft.sh"),
    ("SBI Bank", "SBI-PROD-01", "RTGS Service", "/usr/local/scripts/sbi/rtgs.sh"),
    ("SBI Bank", "SBI-PROD-02", "UPI Gateway", "/usr/local/scripts/sbi/upi_gw.sh"),
    ("SBI Bank", "SBI-PROD-02", "SMS Alerts", "/usr/local/scripts/sbi/sms_alerts.sh"),
    ("SBI Bank", "SBI-DR-01", "Disaster Recovery", "/usr/local/scripts/sbi/dr_sync.sh"),
    ("ICICI Bank", "ICICI-PROD-01", "iMobile", "/usr/local/scripts/icici/imobile.sh"),
    ("ICICI Bank", "ICICI-PROD-01", "Internet Banking", "/usr/local/scripts/icici/netbanking.sh"),
    ("ICICI Bank", "ICICI-PROD-02", "Card Services", "/usr/local/scripts/icici/cards.sh"),
    ("ICICI Bank", "ICICI-PROD-02", "Loan Engine", "/usr/local/scripts/icici/loans.sh"),
    ("Axis Bank", "AXIS-PROD-01", "Mobile App", "/usr/local/scripts/axis/mobile.sh"),
    ("Axis Bank", "AXIS-PROD-01", "IMPS Service", "/usr/local/scripts/axis/imps.sh"),
    ("Axis Bank", "AXIS-PROD-02", "Wealth Management", "/usr/local/scripts/axis/wealth.sh"),
    ("Canara Bank", "CANARA-PROD-01", "Core Banking", "/usr/local/scripts/canara/core.sh"),
    ("Canara Bank", "CANARA-PROD-01", "Net Banking", "/usr/local/scripts/canara/netbanking.sh"),
]

for row, (bank, server, app, path) in enumerate(data, 2):
    ws.cell(row=row, column=1, value=bank)
    ws.cell(row=row, column=2, value=server)
    ws.cell(row=row, column=3, value=app)
    ws.cell(row=row, column=4, value=path)

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 50

wb.save("servers.xlsx")
print("✓ Created servers.xlsx with sample data")
