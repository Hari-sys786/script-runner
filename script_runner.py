#!/usr/bin/env python3
"""
Script Runner v1.0 (Python Edition)
Dark-themed GUI — reads scripts.xlsx, cascading dropdowns, executes bash scripts.
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import subprocess
import threading
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook


# --- Config ---
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts.xlsx")

# --- Dark Theme Colors ---
BG = "#1e1e2e"
BG_SECONDARY = "#2a2a3c"
FG = "#cdd6f4"
FG_DIM = "#6c7086"
ACCENT = "#89b4fa"
ACCENT_HOVER = "#74c7ec"
SUCCESS = "#a6e3a1"
ERROR = "#f38ba8"
WARNING = "#fab387"
BORDER = "#45475a"
INPUT_BG = "#313244"
BUTTON_BG = "#585b70"


class ScriptRunner:
    def __init__(self, root):
        self.root = root
        self.root.title("⚡ Script Runner v1.0")
        self.root.geometry("750x600")
        self.root.configure(bg=BG)
        self.root.minsize(600, 500)

        self.data = {}  # {category: {action: script_path}}
        self.process = None

        self._setup_styles()
        self._build_ui()
        self._load_excel()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Dark.TFrame", background=BG)
        style.configure("Card.TFrame", background=BG_SECONDARY)
        style.configure("Dark.TLabel", background=BG, foreground=FG, font=("Segoe UI", 11))
        style.configure("Title.TLabel", background=BG, foreground=ACCENT, font=("Segoe UI", 18, "bold"))
        style.configure("Subtitle.TLabel", background=BG, foreground=FG_DIM, font=("Segoe UI", 9))
        style.configure("Card.TLabel", background=BG_SECONDARY, foreground=FG, font=("Segoe UI", 11))
        style.configure("Status.TLabel", background=BG, foreground=SUCCESS, font=("Segoe UI", 10))

        style.configure("Dark.TCombobox",
                        fieldbackground=INPUT_BG, background=INPUT_BG,
                        foreground=FG, arrowcolor=ACCENT,
                        borderwidth=1, relief="flat")
        style.map("Dark.TCombobox",
                  fieldbackground=[("readonly", INPUT_BG)],
                  foreground=[("readonly", FG)])

        style.configure("Exec.TButton",
                        background=ACCENT, foreground="#1e1e2e",
                        font=("Segoe UI", 11, "bold"),
                        borderwidth=0, padding=(20, 10))
        style.map("Exec.TButton",
                  background=[("active", ACCENT_HOVER), ("disabled", BORDER)],
                  foreground=[("disabled", FG_DIM)])

        style.configure("Secondary.TButton",
                        background=BUTTON_BG, foreground=FG,
                        font=("Segoe UI", 10),
                        borderwidth=0, padding=(12, 6))
        style.map("Secondary.TButton",
                  background=[("active", BORDER)])

        style.configure("Kill.TButton",
                        background=ERROR, foreground="#1e1e2e",
                        font=("Segoe UI", 10, "bold"),
                        borderwidth=0, padding=(12, 6))
        style.map("Kill.TButton",
                  background=[("active", "#eba0ac")])

    def _build_ui(self):
        # Main container
        main = ttk.Frame(self.root, style="Dark.TFrame")
        main.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Header
        ttk.Label(main, text="⚡ Script Runner", style="Title.TLabel").pack(anchor="w")
        ttk.Label(main, text="Select category → action → execute", style="Subtitle.TLabel").pack(anchor="w", pady=(0, 15))

        # Selection card
        card = ttk.Frame(main, style="Card.TFrame")
        card.pack(fill=tk.X, pady=(0, 10))
        card_inner = ttk.Frame(card, style="Card.TFrame")
        card_inner.pack(fill=tk.X, padx=15, pady=15)

        # Category
        ttk.Label(card_inner, text="Category", style="Card.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 5))
        self.cat_var = tk.StringVar()
        self.cat_combo = ttk.Combobox(card_inner, textvariable=self.cat_var, state="readonly",
                                       style="Dark.TCombobox", width=35, font=("Segoe UI", 11))
        self.cat_combo.grid(row=1, column=0, sticky="ew", padx=(0, 15))
        self.cat_combo.bind("<<ComboboxSelected>>", self._on_category_change)

        # Action
        ttk.Label(card_inner, text="Action", style="Card.TLabel").grid(row=0, column=1, sticky="w", pady=(0, 5))
        self.act_var = tk.StringVar()
        self.act_combo = ttk.Combobox(card_inner, textvariable=self.act_var, state="readonly",
                                       style="Dark.TCombobox", width=35, font=("Segoe UI", 11))
        self.act_combo.grid(row=1, column=1, sticky="ew")

        card_inner.columnconfigure(0, weight=1)
        card_inner.columnconfigure(1, weight=1)

        # Script path display
        self.path_var = tk.StringVar(value="")
        self.path_label = ttk.Label(card_inner, textvariable=self.path_var, style="Card.TLabel",
                                     foreground=FG_DIM, font=("Consolas", 9))
        self.path_label.grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))
        self.act_combo.bind("<<ComboboxSelected>>", self._on_action_change)

        # Buttons row
        btn_frame = ttk.Frame(main, style="Dark.TFrame")
        btn_frame.pack(fill=tk.X, pady=(5, 10))

        self.exec_btn = ttk.Button(btn_frame, text="▶  Execute", style="Exec.TButton", command=self._execute)
        self.exec_btn.pack(side=tk.LEFT)

        self.kill_btn = ttk.Button(btn_frame, text="■ Stop", style="Kill.TButton", command=self._kill)
        self.kill_btn.pack(side=tk.LEFT, padx=(10, 0))
        self.kill_btn.state(["disabled"])

        self.reload_btn = ttk.Button(btn_frame, text="↻ Reload Excel", style="Secondary.TButton", command=self._load_excel)
        self.reload_btn.pack(side=tk.RIGHT)

        self.clear_btn = ttk.Button(btn_frame, text="Clear", style="Secondary.TButton", command=self._clear_output)
        self.clear_btn.pack(side=tk.RIGHT, padx=(0, 10))

        # Output
        ttk.Label(main, text="Output", style="Dark.TLabel").pack(anchor="w", pady=(0, 5))
        self.output = scrolledtext.ScrolledText(main, wrap=tk.WORD, font=("Consolas", 10),
                                                 bg=INPUT_BG, fg=FG, insertbackground=FG,
                                                 relief="flat", borderwidth=0, height=15)
        self.output.pack(fill=tk.BOTH, expand=True)
        self.output.configure(state="disabled")

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        self.status_label = ttk.Label(main, textvariable=self.status_var, style="Status.TLabel")
        self.status_label.pack(anchor="w", pady=(5, 0))

    def _load_excel(self):
        self.data = {}
        if not os.path.exists(EXCEL_FILE):
            self._set_status(f"⚠ Excel not found: {EXCEL_FILE}", WARNING)
            messagebox.showwarning("File Not Found",
                                   f"Create '{EXCEL_FILE}' with columns:\nCategory | Action | Script Path")
            return

        try:
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            for row in rows:
                if not row or len(row) < 3 or not row[0]:
                    continue
                cat, action, path = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
                if cat not in self.data:
                    self.data[cat] = {}
                self.data[cat][action] = path

            categories = sorted(self.data.keys())
            self.cat_combo["values"] = categories
            self.cat_var.set("")
            self.act_combo["values"] = []
            self.act_var.set("")
            self.path_var.set("")

            total = sum(len(v) for v in self.data.values())
            self._set_status(f"✓ Loaded {len(categories)} categories, {total} scripts", SUCCESS)

        except Exception as e:
            self._set_status(f"✗ Error loading Excel: {e}", ERROR)
            messagebox.showerror("Error", str(e))

    def _on_category_change(self, event=None):
        cat = self.cat_var.get()
        if cat in self.data:
            actions = sorted(self.data[cat].keys())
            self.act_combo["values"] = actions
            self.act_var.set("")
            self.path_var.set("")

    def _on_action_change(self, event=None):
        cat = self.cat_var.get()
        act = self.act_var.get()
        if cat in self.data and act in self.data[cat]:
            self.path_var.set(f"📄 {self.data[cat][act]}")

    def _execute(self):
        cat = self.cat_var.get()
        act = self.act_var.get()

        if not cat or not act:
            messagebox.showinfo("Select", "Pick a category and action first.")
            return

        script_path = self.data.get(cat, {}).get(act, "")
        if not script_path:
            messagebox.showerror("Error", "No script path found.")
            return

        if not os.path.exists(script_path):
            self._append_output(f"⚠ Script not found: {script_path}\n", ERROR)
            self._set_status(f"✗ Script not found", ERROR)
            return

        self._append_output(f"\n{'─'*50}\n▶ Running: {act} ({script_path})\n{'─'*50}\n", ACCENT)
        self._set_status(f"⏳ Running: {act}...", WARNING)
        self.exec_btn.state(["disabled"])
        self.kill_btn.state(["!disabled"])

        def run():
            try:
                self.process = subprocess.Popen(
                    ["bash", script_path],
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, bufsize=1
                )
                for line in self.process.stdout:
                    self.root.after(0, self._append_output, line)

                self.process.wait()
                code = self.process.returncode
                if code == 0:
                    self.root.after(0, self._append_output, f"\n✓ Completed (exit 0)\n", SUCCESS)
                    self.root.after(0, self._set_status, f"✓ {act} completed", SUCCESS)
                elif code == -9 or code == -15:
                    self.root.after(0, self._append_output, f"\n■ Killed by user\n", WARNING)
                    self.root.after(0, self._set_status, f"■ Stopped", WARNING)
                else:
                    self.root.after(0, self._append_output, f"\n✗ Failed (exit {code})\n", ERROR)
                    self.root.after(0, self._set_status, f"✗ {act} failed (exit {code})", ERROR)
            except Exception as e:
                self.root.after(0, self._append_output, f"\n✗ Error: {e}\n", ERROR)
                self.root.after(0, self._set_status, f"✗ Error: {e}", ERROR)
            finally:
                self.process = None
                self.root.after(0, lambda: self.exec_btn.state(["!disabled"]))
                self.root.after(0, lambda: self.kill_btn.state(["disabled"]))

        threading.Thread(target=run, daemon=True).start()

    def _kill(self):
        if self.process:
            self.process.kill()

    def _append_output(self, text, color=None):
        self.output.configure(state="normal")
        if color:
            tag = f"color_{color}"
            self.output.tag_configure(tag, foreground=color)
            self.output.insert(tk.END, text, tag)
        else:
            self.output.insert(tk.END, text)
        self.output.see(tk.END)
        self.output.configure(state="disabled")

    def _clear_output(self):
        self.output.configure(state="normal")
        self.output.delete("1.0", tk.END)
        self.output.configure(state="disabled")

    def _set_status(self, text, color=SUCCESS):
        self.status_var.set(text)
        self.status_label.configure(foreground=color)


def main():
    root = tk.Tk()
    root.option_add("*TCombobox*Listbox.background", INPUT_BG)
    root.option_add("*TCombobox*Listbox.foreground", FG)
    root.option_add("*TCombobox*Listbox.selectBackground", ACCENT)
    root.option_add("*TCombobox*Listbox.selectForeground", "#1e1e2e")

    app = ScriptRunner(root)
    root.mainloop()


if __name__ == "__main__":
    main()
