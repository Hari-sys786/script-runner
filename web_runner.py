#!/usr/bin/env python3
"""
Script Runner v1.0 — Web Edition
Same dark UI as desktop version, runs as localhost webpage.
"""

import os
import sys
import json
import subprocess
import threading
import time

try:
    from flask import Flask, render_template_string, request, jsonify, Response
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "flask", "--break-system-packages"])
    from flask import Flask, render_template_string, request, jsonify, Response

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
    from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts.xlsx")

# Track running processes
running_processes = {}
process_lock = threading.Lock()


def load_excel():
    data = {}
    if not os.path.exists(EXCEL_FILE):
        return data
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3 or not row[0]:
            continue
        cat, action, path = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
        if cat not in data:
            data[cat] = {}
        data[cat][action] = path
    wb.close()
    return data


HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>⚡ Script Runner v1.0</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }

        :root {
            --bg: #1e1e2e;
            --bg-secondary: #2a2a3c;
            --fg: #cdd6f4;
            --fg-dim: #6c7086;
            --accent: #89b4fa;
            --accent-hover: #74c7ec;
            --success: #a6e3a1;
            --error: #f38ba8;
            --warning: #fab387;
            --border: #45475a;
            --input-bg: #313244;
            --button-bg: #585b70;
        }

        body {
            background: var(--bg);
            color: var(--fg);
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
            min-height: 100vh;
            padding: 30px;
        }

        .container {
            max-width: 800px;
            margin: 0 auto;
        }

        .header h1 {
            font-size: 28px;
            color: var(--accent);
            margin-bottom: 4px;
        }

        .header p {
            color: var(--fg-dim);
            font-size: 14px;
            margin-bottom: 20px;
        }

        .card {
            background: var(--bg-secondary);
            border-radius: 12px;
            padding: 24px;
            margin-bottom: 16px;
        }

        .form-row {
            display: flex;
            gap: 16px;
            margin-bottom: 12px;
        }

        .form-group {
            flex: 1;
        }

        .form-group label {
            display: block;
            font-size: 13px;
            color: var(--fg-dim);
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        select {
            width: 100%;
            padding: 10px 14px;
            background: var(--input-bg);
            color: var(--fg);
            border: 1px solid var(--border);
            border-radius: 8px;
            font-size: 14px;
            font-family: inherit;
            cursor: pointer;
            appearance: none;
            -webkit-appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%2389b4fa' d='M6 8L1 3h10z'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 12px center;
        }

        select:focus {
            outline: none;
            border-color: var(--accent);
            box-shadow: 0 0 0 2px rgba(137, 180, 250, 0.2);
        }

        select option {
            background: var(--input-bg);
            color: var(--fg);
        }

        .script-path {
            font-family: 'Consolas', 'Fira Code', monospace;
            font-size: 12px;
            color: var(--fg-dim);
            padding: 8px 0 0;
        }

        .btn-row {
            display: flex;
            gap: 10px;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }

        .btn {
            padding: 10px 24px;
            border: none;
            border-radius: 8px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
            font-family: inherit;
        }

        .btn:active { transform: scale(0.97); }

        .btn-exec {
            background: var(--accent);
            color: var(--bg);
        }
        .btn-exec:hover { background: var(--accent-hover); }
        .btn-exec:disabled {
            background: var(--border);
            color: var(--fg-dim);
            cursor: not-allowed;
        }

        .btn-kill {
            background: var(--error);
            color: var(--bg);
        }
        .btn-kill:hover { background: #eba0ac; }
        .btn-kill:disabled {
            background: var(--border);
            color: var(--fg-dim);
            cursor: not-allowed;
        }

        .btn-secondary {
            background: var(--button-bg);
            color: var(--fg);
        }
        .btn-secondary:hover { background: var(--border); }

        .btn-row .spacer { flex: 1; }

        .output-label {
            font-size: 13px;
            color: var(--fg-dim);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }

        #output {
            background: var(--input-bg);
            border: 1px solid var(--border);
            border-radius: 8px;
            padding: 16px;
            font-family: 'Consolas', 'Fira Code', monospace;
            font-size: 13px;
            line-height: 1.6;
            min-height: 300px;
            max-height: 500px;
            overflow-y: auto;
            white-space: pre-wrap;
            word-wrap: break-word;
        }

        #output::-webkit-scrollbar { width: 6px; }
        #output::-webkit-scrollbar-track { background: var(--input-bg); }
        #output::-webkit-scrollbar-thumb { background: var(--border); border-radius: 3px; }

        .status-bar {
            margin-top: 12px;
            font-size: 13px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .status-dot {
            width: 8px;
            height: 8px;
            border-radius: 50%;
            display: inline-block;
        }
        .status-dot.ready { background: var(--success); }
        .status-dot.running { background: var(--warning); animation: pulse 1s infinite; }
        .status-dot.error { background: var(--error); }

        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.4; }
        }

        .text-success { color: var(--success); }
        .text-error { color: var(--error); }
        .text-warning { color: var(--warning); }
        .text-accent { color: var(--accent); }

        @media (max-width: 600px) {
            body { padding: 16px; }
            .form-row { flex-direction: column; }
            .btn-row { flex-direction: column; }
            .btn { width: 100%; text-align: center; }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>⚡ Script Runner</h1>
            <p>Select category → action → execute</p>
        </div>

        <div class="card">
            <div class="form-row">
                <div class="form-group">
                    <label>Category</label>
                    <select id="category" onchange="onCategoryChange()">
                        <option value="">-- Select Category --</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Action</label>
                    <select id="action" onchange="onActionChange()">
                        <option value="">-- Select Action --</option>
                    </select>
                </div>
            </div>
            <div class="script-path" id="scriptPath"></div>
        </div>

        <div class="btn-row">
            <button class="btn btn-exec" id="execBtn" onclick="executeScript()">▶  Execute</button>
            <button class="btn btn-kill" id="killBtn" onclick="killScript()" disabled>■  Stop</button>
            <div class="spacer"></div>
            <button class="btn btn-secondary" onclick="clearOutput()">Clear</button>
            <button class="btn btn-secondary" onclick="reloadExcel()">↻  Reload</button>
        </div>

        <div class="output-label">Output</div>
        <div id="output"></div>

        <div class="status-bar">
            <span class="status-dot ready" id="statusDot"></span>
            <span id="statusText">Ready</span>
        </div>
    </div>

    <script>
        let scriptData = {};
        let currentProcessId = null;
        let eventSource = null;

        // Load data on page load
        fetch('/api/data').then(r => r.json()).then(data => {
            scriptData = data;
            const catSelect = document.getElementById('category');
            Object.keys(data).sort().forEach(cat => {
                const opt = document.createElement('option');
                opt.value = cat;
                opt.textContent = cat;
                catSelect.appendChild(opt);
            });
            setStatus('ready', `Loaded ${Object.keys(data).length} categories`);
        });

        function onCategoryChange() {
            const cat = document.getElementById('category').value;
            const actSelect = document.getElementById('action');
            actSelect.innerHTML = '<option value="">-- Select Action --</option>';
            document.getElementById('scriptPath').textContent = '';

            if (cat && scriptData[cat]) {
                Object.keys(scriptData[cat]).sort().forEach(act => {
                    const opt = document.createElement('option');
                    opt.value = act;
                    opt.textContent = act;
                    actSelect.appendChild(opt);
                });
            }
        }

        function onActionChange() {
            const cat = document.getElementById('category').value;
            const act = document.getElementById('action').value;
            if (cat && act && scriptData[cat] && scriptData[cat][act]) {
                document.getElementById('scriptPath').textContent = '📄 ' + scriptData[cat][act];
            } else {
                document.getElementById('scriptPath').textContent = '';
            }
        }

        function executeScript() {
            const cat = document.getElementById('category').value;
            const act = document.getElementById('action').value;
            if (!cat || !act) {
                alert('Pick a category and action first.');
                return;
            }

            const scriptPath = scriptData[cat][act];
            appendOutput('\\n' + '─'.repeat(50) + '\\n', 'accent');
            appendOutput('▶ Running: ' + act + ' (' + scriptPath + ')\\n', 'accent');
            appendOutput('─'.repeat(50) + '\\n', 'accent');

            setStatus('running', 'Running: ' + act + '...');
            document.getElementById('execBtn').disabled = true;
            document.getElementById('killBtn').disabled = false;

            fetch('/api/execute', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({category: cat, action: act})
            }).then(r => r.json()).then(data => {
                if (data.error) {
                    appendOutput('✗ ' + data.error + '\\n', 'error');
                    setStatus('error', data.error);
                    document.getElementById('execBtn').disabled = false;
                    document.getElementById('killBtn').disabled = true;
                    return;
                }
                currentProcessId = data.process_id;
                pollOutput();
            });
        }

        function pollOutput() {
            if (!currentProcessId) return;

            fetch('/api/output/' + currentProcessId).then(r => r.json()).then(data => {
                if (data.lines && data.lines.length > 0) {
                    data.lines.forEach(line => appendOutput(line));
                }

                if (data.running) {
                    setTimeout(pollOutput, 300);
                } else {
                    if (data.exit_code === 0) {
                        appendOutput('\\n✓ Completed (exit 0)\\n', 'success');
                        setStatus('ready', data.action + ' completed');
                    } else if (data.exit_code === -9 || data.exit_code === -15) {
                        appendOutput('\\n■ Killed by user\\n', 'warning');
                        setStatus('ready', 'Stopped');
                    } else {
                        appendOutput('\\n✗ Failed (exit ' + data.exit_code + ')\\n', 'error');
                        setStatus('error', data.action + ' failed (exit ' + data.exit_code + ')');
                    }
                    currentProcessId = null;
                    document.getElementById('execBtn').disabled = false;
                    document.getElementById('killBtn').disabled = true;
                }
            });
        }

        function killScript() {
            if (currentProcessId) {
                fetch('/api/kill/' + currentProcessId, {method: 'POST'});
            }
        }

        function clearOutput() {
            document.getElementById('output').innerHTML = '';
        }

        function reloadExcel() {
            fetch('/api/reload').then(r => r.json()).then(data => {
                scriptData = data.data;
                const catSelect = document.getElementById('category');
                catSelect.innerHTML = '<option value="">-- Select Category --</option>';
                document.getElementById('action').innerHTML = '<option value="">-- Select Action --</option>';
                document.getElementById('scriptPath').textContent = '';
                Object.keys(data.data).sort().forEach(cat => {
                    const opt = document.createElement('option');
                    opt.value = cat;
                    opt.textContent = cat;
                    catSelect.appendChild(opt);
                });
                setStatus('ready', 'Reloaded: ' + data.categories + ' categories, ' + data.total + ' scripts');
            });
        }

        function appendOutput(text, color) {
            const output = document.getElementById('output');
            const span = document.createElement('span');
            span.textContent = text;
            if (color) span.className = 'text-' + color;
            output.appendChild(span);
            output.scrollTop = output.scrollHeight;
        }

        function setStatus(type, text) {
            const dot = document.getElementById('statusDot');
            const statusText = document.getElementById('statusText');
            dot.className = 'status-dot ' + type;
            statusText.textContent = text;
            statusText.className = type === 'error' ? 'text-error' : type === 'running' ? 'text-warning' : 'text-success';
        }
    </script>
</body>
</html>
"""


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/api/data')
def get_data():
    return jsonify(load_excel())


@app.route('/api/reload')
def reload_data():
    data = load_excel()
    total = sum(len(v) for v in data.values())
    return jsonify({"data": data, "categories": len(data), "total": total})


@app.route('/api/execute', methods=['POST'])
def execute():
    req = request.json
    cat = req.get('category', '')
    act = req.get('action', '')
    data = load_excel()

    if cat not in data or act not in data[cat]:
        return jsonify({"error": "Invalid category or action"})

    script_path = data[cat][act]
    if not os.path.exists(script_path):
        return jsonify({"error": f"Script not found: {script_path}"})

    pid = str(int(time.time() * 1000))

    def run_process():
        try:
            proc = subprocess.Popen(
                ["bash", script_path],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1
            )
            with process_lock:
                running_processes[pid] = {
                    "process": proc,
                    "output": [],
                    "read_index": 0,
                    "running": True,
                    "exit_code": None,
                    "action": act
                }

            for line in proc.stdout:
                with process_lock:
                    if pid in running_processes:
                        running_processes[pid]["output"].append(line)

            proc.wait()
            with process_lock:
                if pid in running_processes:
                    running_processes[pid]["running"] = False
                    running_processes[pid]["exit_code"] = proc.returncode
        except Exception as e:
            with process_lock:
                if pid in running_processes:
                    running_processes[pid]["output"].append(f"Error: {e}\n")
                    running_processes[pid]["running"] = False
                    running_processes[pid]["exit_code"] = 1

    threading.Thread(target=run_process, daemon=True).start()
    time.sleep(0.1)
    return jsonify({"process_id": pid})


@app.route('/api/output/<pid>')
def get_output(pid):
    with process_lock:
        proc_info = running_processes.get(pid)
        if not proc_info:
            return jsonify({"error": "Process not found", "running": False})

        idx = proc_info["read_index"]
        new_lines = proc_info["output"][idx:]
        proc_info["read_index"] = len(proc_info["output"])

        return jsonify({
            "lines": new_lines,
            "running": proc_info["running"],
            "exit_code": proc_info["exit_code"],
            "action": proc_info["action"]
        })


@app.route('/api/kill/<pid>', methods=['POST'])
def kill_process(pid):
    with process_lock:
        proc_info = running_processes.get(pid)
        if proc_info and proc_info["running"]:
            proc_info["process"].kill()
            return jsonify({"status": "killed"})
    return jsonify({"status": "not found"})


if __name__ == '__main__':
    print("⚡ Script Runner Web — http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=False)
